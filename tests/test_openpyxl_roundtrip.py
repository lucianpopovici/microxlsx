"""Validate microxlsx output against openpyxl, an independent reader.

Baselines are authored *by openpyxl* (so they are known-valid, complete
packages with absolute relationship targets, styles, theme, docProps, etc.),
modified with microxlsx, then reopened with openpyxl to confirm the result is
still a valid workbook and the change took effect.
"""
import datetime
import struct
import zlib

import openpyxl
from openpyxl.worksheet.table import Table

from microxlsx.core import XLSXPackage


def _baseline(path, *, with_table=True, with_second_sheet=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row, (name, amount) in enumerate(
            [("Name", "Amount"), ("ada", 10), ("bob", 20)], start=1):
        ws.cell(row, 1, name)
        ws.cell(row, 2, amount)
    ws["D1"] = "=SUM(B2:B3)"
    if with_table:
        ws.add_table(Table(displayName="Sales", ref="A1:B3"))
    if with_second_sheet:
        wb.create_sheet("Data")["A1"] = 99
    wb.save(path)
    return path


def _apply(tmp_path, operation, **baseline_kwargs):
    """Build a baseline, apply ``operation``, save, and reopen with openpyxl."""
    src = _baseline(str(tmp_path / "in.xlsx"), **baseline_kwargs)
    pkg = XLSXPackage(src)
    operation(pkg)
    out = str(tmp_path / "out.xlsx")
    pkg.save(out)
    workbook = openpyxl.load_workbook(out)
    # Force a full parse of every worksheet and its tables.
    for worksheet in workbook.worksheets:
        list(worksheet.iter_rows())
        dict(worksheet.tables)
    return workbook


class TestOpensRealWorkbooks:
    def test_absolute_rel_targets_resolve(self, tmp_path):
        # openpyxl writes Target="/xl/worksheets/sheetN.xml"; the maps must
        # resolve them rather than producing "xl//xl/...".
        pkg = XLSXPackage(_baseline(str(tmp_path / "b.xlsx")))
        assert pkg.sheet_names() == ["Sheet1", "Data"]
        assert pkg.sheet_map["Sheet1"] == "xl/worksheets/sheet1.xml"
        assert pkg.table_names() == ["Sales"]

    def test_read_values_from_openpyxl_file(self, tmp_path):
        pkg = XLSXPackage(_baseline(str(tmp_path / "b.xlsx")))
        assert pkg.get_cell("Sheet1", "A2") == "ada"
        assert pkg.get_cell("Sheet1", "B3") == 20


class TestValueRoundTrips:
    def test_update_cell(self, tmp_path):
        wb = _apply(tmp_path, lambda p: (
            p.update_cell("Sheet1", "B2", value=42),
            p.update_cell("Sheet1", "C1", formula="=B2*2")))
        assert wb["Sheet1"]["B2"].value == 42
        assert wb["Sheet1"]["C1"].value == "=B2*2"

    def test_bool(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.update_cell("Sheet1", "C2", value=True))
        assert wb["Sheet1"]["C2"].value is True

    def test_date(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.update_cell(
            "Sheet1", "E1", value=datetime.date(2024, 1, 15)))
        assert wb["Sheet1"]["E1"].value == datetime.datetime(2024, 1, 15)
        assert wb["Sheet1"]["E1"].is_date

    def test_write_range(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.write_range(
            "Sheet1", "F1", [["x", 1], ["y", 2]]))
        assert [c.value for c in wb["Sheet1"]["F1:G2"][0]] == ["x", 1]
        assert [c.value for c in wb["Sheet1"]["F1:G2"][1]] == ["y", 2]


class TestStyleRoundTrips:
    def test_composed_style(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.update_cell(
            "Sheet1", "A1", value="H",
            style_id=p.add_style(bold=True, fill_color="DDEBF7", border="thin")))
        cell = wb["Sheet1"]["A1"]
        assert cell.font.bold is True
        assert cell.fill.fgColor.rgb == "FFDDEBF7"
        assert cell.border.left.style == "thin"

    def test_number_format(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.update_cell(
            "Sheet1", "B2", value=1234.5,
            style_id=p.add_number_format("$#,##0.00")))
        assert wb["Sheet1"]["B2"].number_format == "$#,##0.00"


class TestTableRoundTrips:
    def test_resize_table(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.resize_table("Sales", add_rows=2))
        assert dict(wb["Sheet1"].tables)["Sales"].ref == "A1:B5"

    def test_append_table_row(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.append_table_row(
            "Sales", {"Name": "cara", "Amount": 30}))
        assert dict(wb["Sheet1"].tables)["Sales"].ref == "A1:B4"
        assert wb["Sheet1"]["A4"].value == "cara"

    def test_add_table(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_table(
            "Data", "Extra", "A1:B3", ["x", "y"]))
        assert "Extra" in dict(wb["Data"].tables)

    def test_remove_table(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.remove_table("Sales"))
        assert not dict(wb["Sheet1"].tables)


class TestStructuralRoundTrips:
    def test_add_sheet(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_sheet("Extra"))
        assert "Extra" in wb.sheetnames

    def test_add_sheet_then_write(self, tmp_path):
        def op(pkg):
            pkg.add_sheet("Extra")
            pkg.update_cell("Extra", "A1", value="hi")
        wb = _apply(tmp_path, op)
        assert wb["Extra"]["A1"].value == "hi"

    def test_remove_sheet(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.remove_sheet("Data"))
        assert wb.sheetnames == ["Sheet1"]

    def test_remove_sheet_with_table(self, tmp_path):
        def op(pkg):
            pkg.add_table("Data", "Extra", "A1:B2", ["x", "y"])
            pkg.remove_sheet("Data")
        wb = _apply(tmp_path, op)
        assert wb.sheetnames == ["Sheet1"]


class TestInsertDeleteRoundTrips:
    def test_insert_rows(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.insert_rows("Sheet1", 2, 2))
        assert wb["Sheet1"]["A4"].value == "ada"  # was A2
        assert dict(wb["Sheet1"].tables)["Sales"].ref == "A1:B5"

    def test_delete_rows(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.delete_rows("Sheet1", 3, 1))
        assert wb["Sheet1"]["A2"].value == "ada"

    def test_insert_cols(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.insert_cols("Sheet1", "C", 1))
        assert wb["Sheet1"]["A1"].value == "Name"

    def test_delete_cols(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.delete_cols("Sheet1", "F", 1))
        assert wb["Sheet1"]["B2"].value == 10


class TestLayoutRoundTrips:
    def test_column_width_and_row_height(self, tmp_path):
        def op(pkg):
            pkg.set_column_width("Sheet1", "B", 25)
            pkg.set_row_height("Sheet1", 1, 30)
        wb = _apply(tmp_path, op)
        assert wb["Sheet1"].column_dimensions["B"].width == 25
        assert wb["Sheet1"].row_dimensions[1].height == 30

    def test_merge_cells(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.merge_cells("Sheet1", "F1:H1"))
        assert "F1:H1" in [str(m) for m in wb["Sheet1"].merged_cells.ranges]


class TestChainedOperations:
    def test_full_report_build(self, tmp_path):
        def op(pkg):
            header = pkg.add_style(bold=True, fill_color="DDEBF7", align="center")
            pkg.add_sheet("Report")
            pkg.write_range("Report", "A1", [["Region", "Q1", "Q2"]], style_id=header)
            pkg.write_range("Report", "A2", [
                ["West", 100, datetime.date(2026, 1, 31)],
                ["East", 200, datetime.date(2026, 2, 28)],
            ])
            pkg.add_table("Report", "Totals", "A1:C3", ["Region", "Q1", "Q2"])
            pkg.set_column_width("Report", "A", 18)
        wb = _apply(tmp_path, op)
        report = wb["Report"]
        assert report["A1"].value == "Region"
        assert report["A1"].font.bold is True
        assert report["C2"].value == datetime.datetime(2026, 1, 31)
        assert "Totals" in dict(report.tables)


class TestFinishingTouchesRoundTrips:
    def test_freeze_panes(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.freeze_panes("Sheet1", "B2"))
        assert wb["Sheet1"].freeze_panes == "B2"

    def test_rename_sheet_updates_cross_sheet_formula(self, tmp_path):
        # Data!A1 references Sheet1; renaming Sheet1 must fix that qualifier.
        def build(path):
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.active["B2"] = 10
            wb.create_sheet("Data")["A1"] = "=Sheet1!B2*2"
            wb.save(path)
            return path
        src = build(str(tmp_path / "in.xlsx"))
        pkg = XLSXPackage(src)
        pkg.rename_sheet("Sheet1", "Sales")
        out = str(tmp_path / "out.xlsx")
        pkg.save(out)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["Sales", "Data"]
        assert wb["Data"]["A1"].value == "=Sales!B2*2"

    def test_add_defined_name(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_defined_name(
            "TaxRate", "Sheet1!$B$2"))
        assert "TaxRate" in wb.defined_names
        assert wb.defined_names["TaxRate"].value == "Sheet1!$B$2"

    def test_add_scoped_defined_name(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_defined_name(
            "Local", "Sheet1!$A$1", sheet_name="Sheet1"))
        assert "Local" in wb["Sheet1"].defined_names

    def test_table_style_reads_back(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_table(
            "Data", "Extra", "A1:B3", ["x", "y"]))
        table = dict(wb["Data"].tables)["Extra"]
        assert table.tableStyleInfo.name == "TableStyleMedium2"
        assert table.tableStyleInfo.showRowStripes is True


class TestTier2RoundTrips:
    def test_worksheet_auto_filter(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.set_auto_filter("Sheet1", "A1:B3"))
        assert wb["Sheet1"].auto_filter.ref == "A1:B3"

    def test_page_setup(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.set_page_setup(
            "Sheet1", orientation="landscape", fit_to_width=1))
        assert wb["Sheet1"].page_setup.orientation == "landscape"
        assert wb["Sheet1"].sheet_properties.pageSetUpPr.fitToPage is True

    def test_print_area(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.set_print_area("Sheet1", "$A$1:$B$3"))
        assert wb["Sheet1"].print_area == "'Sheet1'!$A$1:$B$3"

    def test_header_footer(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.set_header_footer(
            "Sheet1", header="&CQuarterly", footer="&CPage &P of &N"))
        assert wb["Sheet1"].oddHeader.center.text == "Quarterly"

    def test_protect_sheet(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.protect_sheet("Sheet1", password="secret"))
        assert wb["Sheet1"].protection.sheet is True
        assert wb["Sheet1"].protection.password == "DAA7"

    def test_table_autofilter_present(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_table(
            "Data", "Extra", "A1:B3", ["x", "y"]))
        assert dict(wb["Data"].tables)["Extra"].autoFilter.ref == "A1:B3"


def _tiny_png():
    width, height = 4, 2
    raw = b''.join(b'\x00' + b'\x00\x80\xff' * width for _ in range(height))

    def chunk(tag, data):
        return (struct.pack('>I', len(data)) + tag + data
                + struct.pack('>I', zlib.crc32(tag + data) & 0xffffffff))
    ihdr = struct.pack('>IIBBBBB', width, height, 8, 2, 0, 0, 0)
    return (b'\x89PNG\r\n\x1a\n' + chunk(b'IHDR', ihdr)
            + chunk(b'IDAT', zlib.compress(raw)) + chunk(b'IEND', b''))


class TestTier3RoundTrips:
    def test_hyperlink(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_hyperlink(
            "Sheet1", "A1", "https://example.com", tooltip="Go"))
        link = wb["Sheet1"]["A1"].hyperlink
        assert link.target == "https://example.com"
        assert link.tooltip == "Go"

    def test_image(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_image("Sheet1", "B2", _tiny_png()))
        images = wb["Sheet1"]._images  # pylint: disable=protected-access
        assert len(images) == 1
        assert images[0].anchor._from.col == 1  # pylint: disable=protected-access

    def test_comment(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_comment(
            "Sheet1", "A1", "Nice cell", author="Alice"))
        assert wb["Sheet1"]["A1"].comment.text == "Nice cell"
        assert wb["Sheet1"]["A1"].comment.author == "Alice"

    def test_multiple_comments(self, tmp_path):
        def op(pkg):
            pkg.add_comment("Sheet1", "A1", "one", author="A")
            pkg.add_comment("Sheet1", "C3", "two", author="B")
        wb = _apply(tmp_path, op)
        assert wb["Sheet1"]["A1"].comment.text == "one"
        assert wb["Sheet1"]["C3"].comment.text == "two"

    def test_image_and_comment_together(self, tmp_path):
        def op(pkg):
            pkg.add_image("Sheet1", "D4", _tiny_png())
            pkg.add_comment("Sheet1", "A1", "note", author="A")
        wb = _apply(tmp_path, op)
        assert len(wb["Sheet1"]._images) == 1  # pylint: disable=protected-access
        assert wb["Sheet1"]["A1"].comment.text == "note"


class TestTier4RoundTrips:
    def test_list_validation(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_data_validation(
            "Sheet1", "C1:C5", "list", formula1='"Yes,No,Maybe"',
            prompt="Pick", prompt_title="Choose"))
        dvs = list(wb["Sheet1"].data_validations.dataValidation)
        assert len(dvs) == 1
        assert dvs[0].type == "list"
        assert dvs[0].formula1 == '"Yes,No,Maybe"'
        assert str(dvs[0].sqref) == "C1:C5"
        assert dvs[0].promptTitle == "Choose"

    def test_whole_between_validation(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_data_validation(
            "Sheet1", "C1:C5", "whole", operator="between",
            formula1="1", formula2="100", error="1-100", error_title="Bad"))
        dv = list(wb["Sheet1"].data_validations.dataValidation)[0]
        assert dv.type == "whole"
        assert dv.operator == "between"
        assert dv.formula1 == "1"
        assert dv.formula2 == "100"
        assert dv.errorTitle == "Bad"

    def test_two_validations(self, tmp_path):
        def op(pkg):
            pkg.add_data_validation("Sheet1", "C1:C5", "list",
                                    formula1='"a,b"')
            pkg.add_data_validation("Sheet1", "D1:D5", "decimal",
                                    operator="greaterThan", formula1="0")
        wb = _apply(tmp_path, op)
        assert len(list(wb["Sheet1"].data_validations.dataValidation)) == 2

    def test_cellis_rule_with_dxf(self, tmp_path):
        def op(pkg):
            style = pkg.add_dxf(fill_color="FFC7CE", font_color="9C0006")
            pkg.add_conditional_format("Sheet1", "B2:B3", "cellIs",
                                       operator="greaterThan", formula="15",
                                       dxf=style)
        wb = _apply(tmp_path, op)
        rules = wb["Sheet1"].conditional_formatting["B2:B3"]
        assert rules[0].type == "cellIs"
        assert rules[0].operator == "greaterThan"
        assert rules[0].dxfId == 0

    def test_color_scale_rule(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_conditional_format(
            "Sheet1", "B2:B3", "colorScale",
            colors=["FFF8696B", "FFFCFCFF", "FF63BE7B"]))
        rule = wb["Sheet1"].conditional_formatting["B2:B3"][0]
        assert rule.type == "colorScale"
        assert len(rule.colorScale.color) == 3

    def test_data_bar_rule(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.add_conditional_format(
            "Sheet1", "B2:B3", "dataBar", color="638EC6"))
        rule = wb["Sheet1"].conditional_formatting["B2:B3"][0]
        assert rule.type == "dataBar"

    def test_multiple_rules_get_priorities(self, tmp_path):
        def op(pkg):
            pkg.add_conditional_format("Sheet1", "B2:B3", "dataBar")
            pkg.add_conditional_format("Sheet1", "B2:B3", "colorScale")
        wb = _apply(tmp_path, op)
        rules = wb["Sheet1"].conditional_formatting["B2:B3"]
        assert sorted(r.priority for r in rules) == [1, 2]


class TestTier5RoundTrips:
    def test_hide_rows(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.hide_rows("Sheet1", 2, 3))
        assert wb["Sheet1"].row_dimensions[2].hidden is True
        assert wb["Sheet1"].row_dimensions[3].hidden is True

    def test_hide_columns(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.hide_columns("Sheet1", "B"))
        assert wb["Sheet1"].column_dimensions["B"].hidden is True

    def test_group_rows(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.group_rows(
            "Sheet1", 2, 3, collapsed=True))
        assert wb["Sheet1"].row_dimensions[2].outlineLevel == 1
        assert wb["Sheet1"].row_dimensions[2].hidden is True

    def test_group_columns(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.group_columns("Sheet1", "B", "C"))
        assert wb["Sheet1"].column_dimensions["B"].outlineLevel == 1

    def test_tab_color(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.set_tab_color("Sheet1", "FF0000"))
        assert wb["Sheet1"].sheet_properties.tabColor.rgb == "FFFF0000"

    def test_sheet_visibility(self, tmp_path):
        wb = _apply(tmp_path, lambda p: p.set_sheet_visibility("Data", "hidden"))
        assert wb["Data"].sheet_state == "hidden"

    def test_remove_hyperlink(self, tmp_path):
        def op(pkg):
            pkg.add_hyperlink("Sheet1", "A1", "https://example.com")
            pkg.remove_hyperlink("Sheet1", "A1")
        wb = _apply(tmp_path, op)
        assert wb["Sheet1"]["A1"].hyperlink is None

    def test_remove_comment(self, tmp_path):
        def op(pkg):
            pkg.add_comment("Sheet1", "A1", "gone", author="A")
            pkg.add_comment("Sheet1", "C3", "stays", author="A")
            pkg.remove_comment("Sheet1", "A1")
        wb = _apply(tmp_path, op)
        assert wb["Sheet1"]["A1"].comment is None
        assert wb["Sheet1"]["C3"].comment.text == "stays"

    def test_remove_defined_name(self, tmp_path):
        def op(pkg):
            pkg.add_defined_name("Keep", "Sheet1!$A$1")
            pkg.add_defined_name("Drop", "Sheet1!$B$1")
            assert pkg.remove_defined_name("Drop") == 1
        wb = _apply(tmp_path, op)
        assert "Keep" in wb.defined_names
        assert "Drop" not in wb.defined_names
